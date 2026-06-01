from __future__ import annotations

import csv
from pathlib import Path

from .models import DataSource, DVMEdge


def infer_edges_for_datasource(datasource: DataSource, *, base_dir: str | Path = ".") -> list[DVMEdge]:
    if datasource.type == "csv":
        columns = _csv_columns(datasource, Path(base_dir))
    elif datasource.type == "excel":
        columns = _excel_columns(datasource, Path(base_dir))
    else:
        return []

    if len(columns) < 2:
        return []

    root = columns[0]
    return [
        DVMEdge.create(
            root,
            column,
            datasource.name,
            key_positions=1,
            value_positions=index,
            selected=True,
            head_description=f"Inferred key column from {datasource.name}",
            tail_description=f"Inferred value column from {datasource.name}",
            description="Inferred from datasource columns",
        )
        for index, column in enumerate(columns[1:], start=2)
    ]


def _csv_columns(datasource: DataSource, base_dir: Path) -> list[str]:
    file_path = _resolve_file(datasource, base_dir)
    delimiter = datasource.get("delimiter", ",") or ","
    headings = datasource.get("headings", "no").lower() == "yes"
    with file_path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.reader(handle, delimiter=delimiter)
        first_row = next(reader, [])
    return _normalize_columns(first_row, headings=headings)


def _excel_columns(datasource: DataSource, base_dir: Path) -> list[str]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Excel inference requires the optional dependency: pip install datamingler[excel]") from exc

    file_path = _resolve_file(datasource, base_dir)
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    try:
        sheet_name = datasource.get("sheet")
        sheet = workbook[sheet_name] if sheet_name else workbook.active
        first_row = next(sheet.iter_rows(values_only=True), [])
    finally:
        workbook.close()
    headings = datasource.get("headings", "yes").lower() == "yes"
    return _normalize_columns(["" if value is None else str(value) for value in first_row], headings=headings)


def _normalize_columns(row: list[str], *, headings: bool) -> list[str]:
    columns = []
    for index, value in enumerate(row, start=1):
        text = str(value or "").strip() if headings else f"column_{index}"
        columns.append(text or f"column_{index}")
    return columns


def _resolve_file(datasource: DataSource, base_dir: Path) -> Path:
    path = datasource.get("path")
    filename = datasource.get("filename")
    value = str(Path(path) / filename) if path else filename
    candidate = Path(value.strip())
    if candidate.is_absolute() and candidate.exists():
        return candidate
    for item in (base_dir / candidate, Path.cwd() / candidate, candidate):
        if item.exists():
            return item
    raise FileNotFoundError(f"Could not resolve datasource path: {value!r}")
