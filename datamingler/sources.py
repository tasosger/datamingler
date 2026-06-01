from __future__ import annotations

import csv
import shlex
import sqlite3
import subprocess
from contextlib import closing
from pathlib import Path
from typing import Iterable, Sequence

from .graph import DVMGraph
from .kvstore import KeyListStore
from .models import DataSource, DVMEdge
from .xmlio import load_datasources_xml


class DataSourceRegistry:
    def __init__(self, datasources: dict[str, DataSource] | None = None, *, base_dir: str | Path = ".") -> None:
        self.datasources = datasources or {}
        self.base_dir = Path(base_dir)

    @classmethod
    def from_xml(cls, path: str | Path) -> "DataSourceRegistry":
        path = Path(path)
        return cls(load_datasources_xml(path), base_dir=path.parent)

    def require(self, name: str) -> DataSource:
        key = name.strip()
        if key not in self.datasources:
            available = ", ".join(sorted(self.datasources))
            raise KeyError(f"Datasource {key!r} not found. Available: {available}")
        return self.datasources[key]


class EdgeMaterializer:
    def __init__(self, registry: DataSourceRegistry) -> None:
        self.registry = registry

    def materialize_graph_edge(
        self,
        edge: DVMEdge,
        store: KeyListStore,
        *,
        root_alias: str | None = None,
        child_alias: str | None = None,
    ) -> None:
        datasource = self.registry.require(edge.datasource)
        root = root_alias or edge.head_name
        child = child_alias or edge.tail_name
        with store.pipeline() as pipe:
            for row in self.iter_rows(datasource, edge.query):
                key = _join_positions(row, edge.key_positions)
                value = _join_positions(row, edge.value_positions)
                pipe.add_value(root, child, key, value)

    def materialize_pair(
        self,
        graph: DVMGraph,
        head_name: str,
        tail_name: str,
        store: KeyListStore,
        *,
        root_alias: str | None = None,
        child_alias: str | None = None,
    ) -> None:
        edges = graph.require_edge(head_name, tail_name)
        for edge in edges:
            self.materialize_graph_edge(edge, store, root_alias=root_alias, child_alias=child_alias)

    def iter_rows(self, datasource: DataSource, query: str = "") -> Iterable[list[str]]:
        if datasource.type == "csv":
            return self._iter_csv(datasource)
        if datasource.type == "excel":
            return self._iter_excel(datasource)
        if datasource.type == "db":
            return self._iter_db(datasource, query)
        if datasource.type == "process":
            return self._iter_process(datasource)
        raise NotImplementedError(f"Unsupported datasource type: {datasource.type!r}")

    def _iter_csv(self, datasource: DataSource) -> Iterable[list[str]]:
        file_path = self._resolve_file(datasource)
        delimiter = datasource.get("delimiter", ",") or ","
        headings = datasource.get("headings", "no").lower() == "yes"
        with file_path.open(newline="", encoding="utf-8-sig") as handle:
            reader = csv.reader(handle, delimiter=delimiter)
            for index, row in enumerate(reader):
                if index == 0 and headings:
                    continue
                yield [cell.strip() for cell in row]

    def _iter_excel(self, datasource: DataSource) -> Iterable[list[str]]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("Excel datasources require the optional dependency: pip install datamingler[excel]") from exc

        file_path = self._resolve_file(datasource)
        sheet_name = datasource.get("sheet")
        headings = datasource.get("headings", "no").lower() == "yes"
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        try:
            sheet = workbook[sheet_name] if sheet_name else workbook.active
            for index, row in enumerate(sheet.iter_rows(values_only=True)):
                if index == 0 and headings:
                    continue
                yield ["" if value is None else str(value) for value in row]
        finally:
            workbook.close()

    def _iter_db(self, datasource: DataSource, query: str) -> Iterable[list[str]]:
        if not query.strip():
            raise ValueError(f"DB datasource {datasource.name!r} requires an edge query")
        system = datasource.get("system").lower()
        connection = datasource.get("connection")
        if system in {"sqlite", "sqlite3"}:
            db_path = self._resolve_path(connection)
            with closing(sqlite3.connect(db_path)) as conn:
                cursor = conn.execute(query)
                try:
                    for row in cursor:
                        yield ["" if value is None else str(value) for value in row]
                finally:
                    cursor.close()
            return

        if "://" in connection:
            try:
                from sqlalchemy import create_engine, text
            except ImportError as exc:
                raise RuntimeError("SQLAlchemy URL datasources require: pip install datamingler[db]") from exc
            engine = create_engine(connection)
            try:
                with engine.connect() as conn:
                    result = conn.execute(text(query))
                    try:
                        for row in result:
                            yield ["" if value is None else str(value) for value in row]
                    finally:
                        result.close()
            finally:
                engine.dispose()
            return

        raise NotImplementedError(
            f"DB system {system!r} is not supported without a SQLAlchemy URL. "
            "Use system=sqlite or set connection to a SQLAlchemy URL."
        )

    def _iter_process(self, datasource: DataSource) -> Iterable[list[str]]:
        delimiter = datasource.get("delimiter", ",") or ","
        system = datasource.get("system") or datasource.get("command")
        engine = datasource.get("engine")
        engine_path = datasource.get("enginePath")
        filename = datasource.get("filename")

        if system:
            command = shlex.split(system)
        elif engine:
            command = [str(Path(engine_path) / engine) if engine_path else engine]
        else:
            raise ValueError(f"Process datasource {datasource.name!r} requires a system command or engine")

        if filename:
            command.append(str(self._resolve_file(datasource)))

        completed = subprocess.run(command, check=True, capture_output=True, text=True)
        for line in completed.stdout.splitlines():
            if line.strip():
                yield [part.strip() for part in line.split(delimiter)]

    def _resolve_file(self, datasource: DataSource) -> Path:
        path = datasource.get("path")
        filename = datasource.get("filename")
        return self._resolve_path(str(Path(path) / filename) if path else filename)

    def _resolve_path(self, value: str) -> Path:
        candidate = Path(value.strip())
        if candidate.is_absolute() and candidate.exists():
            return candidate
        candidates = [
            self.registry.base_dir / candidate,
            Path.cwd() / candidate,
            candidate,
        ]
        for item in candidates:
            if item.exists():
                return item
        raise FileNotFoundError(f"Could not resolve datasource path: {value!r}")


def _join_positions(row: Sequence[str], positions: tuple[int, ...]) -> str:
    if not positions:
        return ""
    values: list[str] = []
    for position in positions:
        index = position - 1
        if index < 0 or index >= len(row):
            raise IndexError(f"Column position {position} is outside row with {len(row)} columns")
        values.append(row[index])
    return ":".join(values)
