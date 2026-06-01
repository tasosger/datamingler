from __future__ import annotations

import sqlite3
import sys
import tempfile
import unittest
from contextlib import closing
from pathlib import Path

import fakeredis

from datamingler.engine import QueryEvaluator
from datamingler.kvstore import KeyListStore
from datamingler.models import DataSource
from datamingler.sources import DataSourceRegistry, EdgeMaterializer
from datamingler.xmlio import load_dvm_xml, load_query_text, parse_query_text, save_query_xml, load_query_xml


ROOT = Path(__file__).resolve().parents[1]
EXAMPLES = ROOT / "examples"


def _make_store() -> KeyListStore:
    """Return a KeyListStore backed by fakeredis (no live Redis needed in tests)."""
    store = object.__new__(KeyListStore)
    store._redis = fakeredis.FakeRedis(decode_responses=True)
    return store


class DataMinglerCoreTests(unittest.TestCase):
    def test_text_query_round_trip(self) -> None:
        plan = parse_query_text((EXAMPLES / "customer_summary.qdvm").read_text(encoding="utf-8"))
        self.assertEqual(plan.root, "X")
        self.assertEqual(plan.nodes["X"].children, ("A", "G", "N"))
        self.assertTrue(plan.nodes["A"].output)
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "query.xml"
            save_query_xml(plan, path)
            loaded = load_query_xml(path)
        self.assertEqual(loaded.root, plan.root)
        self.assertEqual(loaded.nodes["N"].transformations[0].name, "map")

    def test_customer_summary_evaluation(self) -> None:
        graph = load_dvm_xml(EXAMPLES / "sample.dvm.xml")
        registry = DataSourceRegistry.from_xml(EXAMPLES / "datasources.xml")
        plan = load_query_text(EXAMPLES / "customer_summary.qdvm")
        result = QueryEvaluator(graph, registry, store=_make_store()).evaluate(plan)
        rows = {row["custID-X"]: row for row in result.to_rows()}
        self.assertEqual(rows["C1"]["Age-A"], "35")
        self.assertEqual(rows["C1"]["Gender-G"], "F")
        self.assertEqual(rows["C1"]["Comment-N"], "10.0")
        self.assertEqual(rows["C2"]["Comment-N"], "3.0")

    def test_nested_theta_json_evaluation(self) -> None:
        graph = load_dvm_xml(EXAMPLES / "sample.dvm.xml")
        registry = DataSourceRegistry.from_xml(EXAMPLES / "datasources.xml")
        plan = load_query_text(EXAMPLES / "transactions_after_2019.qdvm")
        result = QueryEvaluator(graph, registry, store=_make_store()).evaluate(plan)
        records = {record["custID-X"]: record for record in result.to_json_records()}
        self.assertEqual(records["C1"]["transID-T_s"], [{"transID-T": "T1", "Amount-Amt": "10"}])
        self.assertEqual(records["C2"]["transID-T_s"], [{"transID-T": "T3", "Amount-Amt": "5"}])


class DataSourceTests(unittest.TestCase):
    def _materializer(self, temp_dir: Path) -> EdgeMaterializer:
        return EdgeMaterializer(DataSourceRegistry(base_dir=temp_dir))

    def test_csv_datasource_reads_rows_and_skips_headings(self) -> None:
        with tempfile.TemporaryDirectory() as temp:
            temp_dir = Path(temp)
            (temp_dir / "customers.csv").write_text("id,age\nC1,35\nC2,41\n", encoding="utf-8")
            datasource = DataSource(
                name="customers_csv",
                type="csv",
                options={"filename": "customers.csv", "headings": "yes"},
            )

            rows = list(self._materializer(temp_dir).iter_rows(datasource))

        self.assertEqual(rows, [["C1", "35"], ["C2", "41"]])

    def test_sqlite_db_datasource_runs_edge_query(self) -> None:
        with tempfile.TemporaryDirectory() as temp:
            temp_dir = Path(temp)
            db_path = temp_dir / "customers.sqlite"
            with closing(sqlite3.connect(db_path)) as conn:
                conn.execute("create table customers (id text, age integer)")
                conn.executemany("insert into customers values (?, ?)", [("C1", 35), ("C2", 41)])
                conn.commit()
            datasource = DataSource(
                name="customers_db",
                type="db",
                options={"system": "sqlite", "connection": "customers.sqlite"},
            )

            rows = list(self._materializer(temp_dir).iter_rows(datasource, "select id, age from customers order by id"))

        self.assertEqual(rows, [["C1", "35"], ["C2", "41"]])

    def test_process_datasource_accepts_frontend_system_command(self) -> None:
        with tempfile.TemporaryDirectory() as temp:
            temp_dir = Path(temp)
            script = temp_dir / "emit_rows.py"
            data = temp_dir / "input.txt"
            script.write_text(
                "from pathlib import Path\n"
                "import sys\n"
                "for line in Path(sys.argv[1]).read_text(encoding='utf-8').splitlines():\n"
                "    print(line.upper())\n",
                encoding="utf-8",
            )
            data.write_text("c1,small\nc2,large\n", encoding="utf-8")
            datasource = DataSource(
                name="customers_process",
                type="process",
                options={
                    "system": f'"{sys.executable}" "{script}"',
                    "filename": "input.txt",
                    "delimiter": ",",
                },
            )

            rows = list(self._materializer(temp_dir).iter_rows(datasource))

        self.assertEqual(rows, [["C1", "SMALL"], ["C2", "LARGE"]])

    def test_process_datasource_keeps_legacy_engine_fields(self) -> None:
        with tempfile.TemporaryDirectory() as temp:
            temp_dir = Path(temp)
            data = temp_dir / "input.py"
            data.write_text("print('C1,35')\nprint('C2,41')\n", encoding="utf-8")
            datasource = DataSource(
                name="customers_process",
                type="process",
                options={"engine": sys.executable, "filename": "input.py", "delimiter": ","},
            )

            rows = list(self._materializer(temp_dir).iter_rows(datasource))

        self.assertEqual(rows, [["C1", "35"], ["C2", "41"]])

    def test_excel_datasource_reads_rows_when_dependency_is_available(self) -> None:
        try:
            from openpyxl import Workbook
        except ImportError:
            self.skipTest("openpyxl is not installed")

        with tempfile.TemporaryDirectory() as temp:
            temp_dir = Path(temp)
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Customers"
            sheet.append(["id", "age"])
            sheet.append(["C1", 35])
            sheet.append(["C2", 41])
            workbook.save(temp_dir / "customers.xlsx")
            datasource = DataSource(
                name="customers_excel",
                type="excel",
                options={"filename": "customers.xlsx", "sheet": "Customers", "headings": "yes"},
            )

            rows = list(self._materializer(temp_dir).iter_rows(datasource))

        self.assertEqual(rows, [["C1", "35"], ["C2", "41"]])


if __name__ == "__main__":
    unittest.main()
